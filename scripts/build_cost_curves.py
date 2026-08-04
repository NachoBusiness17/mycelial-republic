"""Build compute cost-curve analyst workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter

wb = Workbook()

blue = Font(name="Arial", color="0000FF", size=11)
black = Font(name="Arial", color="000000", size=11)
header = Font(name="Arial", bold=True, size=11, color="FFFFFF")
title = Font(name="Arial", bold=True, size=16, color="1A1A2E")
section = Font(name="Arial", bold=True, size=12, color="16213E")
yellow = PatternFill("solid", fgColor="FFFF99")
navy = PatternFill("solid", fgColor="1A1A2E")
cur0 = '$#,##0;($#,##0);"-"'
pct = "0.0%"


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header
        cell.fill = navy
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


# ========== ASSUMPTIONS ==========
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "Sovereign Mirror — Compute Cost Curves"
ws["A1"].font = title
ws.merge_cells("A1:F1")
ws["A2"] = (
    "Analyst model: Local lab vs GPU rental (GSTD-class / cloud). "
    "Blue/yellow = inputs. All other values are formulas."
)
ws["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")
ws.merge_cells("A2:F2")

ws["A4"] = "1. LOCAL LAB CAPEX SCENARIOS"
ws["A4"].font = section
headers = [
    "Scenario",
    "Capex ($)",
    "Annual power/ops ($)",
    "Useful life (yrs)",
    "Annualized capex ($/yr)",
    "Fully loaded yr cost ($)",
]
for i, h in enumerate(headers, 1):
    ws.cell(5, i, h)
style_header_row(ws, 5, 6)

labs = [
    ("Lab_5k", 5000, 250, 4),
    ("Lab_15k", 15000, 400, 4),
    ("Lab_25k", 25000, 550, 4),
]
for i, (name, capex, opex, life) in enumerate(labs):
    r = 6 + i
    ws.cell(r, 1, name).font = black
    c = ws.cell(r, 2, capex)
    c.font = blue
    c.fill = yellow
    c.number_format = cur0
    c = ws.cell(r, 3, opex)
    c.font = blue
    c.fill = yellow
    c.number_format = cur0
    c = ws.cell(r, 4, life)
    c.font = blue
    c.fill = yellow
    ws.cell(r, 5, f"=B{r}/D{r}").number_format = cur0
    ws.cell(r, 6, f"=E{r}+C{r}").number_format = cur0

ws["A10"] = "2. RENTAL COST PER TRAINING RUN ($/job)"
ws["A10"].font = section
for i, h in enumerate(
    ["Job profile", "GSTD / decen. net ($)", "Commodity cloud ($)", "Blended rental ($)", "Notes"],
    1,
):
    ws.cell(11, i, h)
style_header_row(ws, 11, 5)

jobs = [
    ("QLoRA_7B_14B", 15, 25, "Starter mirror iterate"),
    ("QLoRA_32B", 60, 100, "Athena-class scale"),
    ("QLoRA_70B_or_long", 150, 250, "Burst / research"),
    ("Infer_month_heavy", 40, 60, "Always-on rented infer equiv"),
]
for i, (name, gstd, cloud, note) in enumerate(jobs):
    r = 12 + i
    ws.cell(r, 1, name)
    c = ws.cell(r, 2, gstd)
    c.font = blue
    c.fill = yellow
    c.number_format = cur0
    c = ws.cell(r, 3, cloud)
    c.font = blue
    c.fill = yellow
    c.number_format = cur0
    ws.cell(r, 4, f"=ROUND((B{r}+C{r})/2,0)").number_format = cur0
    ws.cell(r, 5, note).font = Font(name="Arial", size=9, color="666666")

ws["A17"] = "3. USAGE INTENSITY (runs / year of primary job)"
ws["A17"].font = section
for i, h in enumerate(["Intensity", "Runs/year", "Description"], 1):
    ws.cell(18, i, h)
style_header_row(ws, 18, 3)
intens = [
    ("Sparse", 4, "Quarterly retrain / experiments"),
    ("Moderate", 24, "Twice monthly iteration"),
    ("Heavy", 52, "Weekly train cycle"),
    ("Industrial", 120, "Near-continuous R&D"),
]
for i, (name, runs, desc) in enumerate(intens):
    r = 19 + i
    ws.cell(r, 1, name)
    c = ws.cell(r, 2, runs)
    c.font = blue
    c.fill = yellow
    ws.cell(r, 3, desc)

ws["A24"] = "4. GLOBAL CONTROLS"
ws["A24"].font = section
ws["A25"] = "Primary job profile row (12=7B, 13=32B, 14=70B)"
c = ws.cell(25, 2, 13)
c.font = blue
c.fill = yellow
ws["A26"] = "Horizon years for TCO"
c = ws.cell(26, 2, 5)
c.font = blue
c.fill = yellow
ws["A27"] = "Discount rate (reference)"
c = ws.cell(27, 2, 0.08)
c.font = blue
c.fill = yellow
c.number_format = pct
ws["A28"] = "Sovereignty premium ($/yr value of local control)"
c = ws.cell(28, 2, 500)
c.font = blue
c.fill = yellow
c.number_format = cur0
ws["A29"] = "Data-risk penalty for rental ($/yr)"
c = ws.cell(29, 2, 300)
c.font = blue
c.fill = yellow
c.number_format = cur0

ws["A31"] = "Sources / methodology"
ws["A31"].font = section
ws["A32"] = (
    "Capex: project lab BOM tiers ($5k/$15k/$25k). Rental: illustrative mid-2026 GPU market "
    "ranges for QLoRA-class jobs (commodity cloud + decentralized nets). NOT an official GSTD "
    "tariff — replace yellow cells with live quotes. Power: ~$0.12–0.20/kWh duty cycle estimate. "
    "Sovereignty/risk adders are explicit judgment parameters for adjusted TCO."
)
ws["A32"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A32:F34")

for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 18 if col > 1 else 44
ws.column_dimensions["E"].width = 36

# ========== BREAKEVEN ==========
ws2 = wb.create_sheet("BreakEven")
ws2["A1"] = "Break-even analysis — rental vs local lab"
ws2["A1"].font = title
ws2.merge_cells("A1:H1")
ws2["A2"] = (
    "Break-even runs = Capex / ($/run). Below that run count, rent wins on cash; above, own wins."
)
ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

ws2["A4"] = "Primary job $/run (blended, from Assumptions)"
ws2["B4"] = '=INDEX(Assumptions!D12:D15,Assumptions!B25-11)'
ws2["B4"].number_format = cur0
ws2["B4"].font = Font(name="Arial", bold=True, size=14)

headers = [
    "Lab scenario",
    "Capex",
    "$/run",
    "Break-even runs",
    "Yrs @ Sparse(4)",
    "Yrs @ Moderate(24)",
    "Yrs @ Heavy(52)",
    "Yrs @ Industrial(120)",
]
for i, h in enumerate(headers, 1):
    ws2.cell(6, i, h)
style_header_row(ws2, 6, 8)

for i in range(3):
    r = 7 + i
    ar = 6 + i
    ws2.cell(r, 1, f"=Assumptions!A{ar}")
    ws2.cell(r, 2, f"=Assumptions!B{ar}").number_format = cur0
    ws2.cell(r, 3, "=$B$4").number_format = cur0
    ws2.cell(r, 4, f"=IF(C{r}=0,0,B{r}/C{r})").number_format = "0.0"
    ws2.cell(r, 5, f"=D{r}/Assumptions!$B$19").number_format = "0.00"
    ws2.cell(r, 6, f"=D{r}/Assumptions!$B$20").number_format = "0.00"
    ws2.cell(r, 7, f"=D{r}/Assumptions!$B$21").number_format = "0.00"
    ws2.cell(r, 8, f"=D{r}/Assumptions!$B$22").number_format = "0.00"

ws2["A11"] = "Sensitivity: break-even runs vs $/run"
ws2["A11"].font = section
for i, h in enumerate(["$/run", "BE Lab_5k", "BE Lab_15k", "BE Lab_25k"], 1):
    ws2.cell(12, i, h)
style_header_row(ws2, 12, 4)

prices = [5, 10, 15, 20, 25, 30, 40, 50, 60, 75, 100, 125, 150, 175, 200]
for i, p in enumerate(prices):
    r = 13 + i
    c = ws2.cell(r, 1, p)
    c.font = blue
    c.fill = yellow
    c.number_format = cur0
    ws2.cell(r, 2, f"=Assumptions!$B$6/A{r}").number_format = "0.0"
    ws2.cell(r, 3, f"=Assumptions!$B$7/A{r}").number_format = "0.0"
    ws2.cell(r, 4, f"=Assumptions!$B$8/A{r}").number_format = "0.0"

chart = LineChart()
chart.title = "Break-even runs vs rental $/job"
chart.style = 10
chart.y_axis.title = "Runs to break even"
chart.x_axis.title = "$ per training run"
chart.height = 12
chart.width = 18
data = Reference(ws2, min_col=2, min_row=12, max_col=4, max_row=12 + len(prices))
cats = Reference(ws2, min_col=1, min_row=13, max_row=12 + len(prices))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, "F11")

for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 14
ws2.column_dimensions["A"].width = 18

# ========== TCO CURVES ==========
ws3 = wb.create_sheet("TCO_Curves")
ws3["A1"] = "5-year cumulative cash TCO curves"
ws3["A1"].font = title
ws3.merge_cells("A1:H1")
ws3["A2"] = (
    "Year 0: lab = full capex, rent = $0. Later years: rent = cumulative jobs; lab = capex + cumulative ops."
)
ws3["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

ws3["A4"] = "Job $/run"
ws3["B4"] = '=INDEX(Assumptions!D12:D15,Assumptions!B25-11)'
ws3["B4"].number_format = cur0
ws3["B4"].font = Font(name="Arial", bold=True)

for i, h in enumerate(
    [
        "Year",
        "Rent Sparse",
        "Rent Moderate",
        "Rent Heavy",
        "Rent Industrial",
        "Own Lab_5k",
        "Own Lab_15k",
        "Own Lab_25k",
    ],
    1,
):
    ws3.cell(6, i, h)
style_header_row(ws3, 6, 8)

for y in range(0, 6):
    r = 7 + y
    ws3.cell(r, 1, y)
    if y == 0:
        for col in range(2, 6):
            ws3.cell(r, col, 0).number_format = cur0
        ws3.cell(r, 6, "=Assumptions!B6").number_format = cur0
        ws3.cell(r, 7, "=Assumptions!B7").number_format = cur0
        ws3.cell(r, 8, "=Assumptions!B8").number_format = cur0
    else:
        ws3.cell(r, 2, f"=A{r}*Assumptions!$B$19*$B$4").number_format = cur0
        ws3.cell(r, 3, f"=A{r}*Assumptions!$B$20*$B$4").number_format = cur0
        ws3.cell(r, 4, f"=A{r}*Assumptions!$B$21*$B$4").number_format = cur0
        ws3.cell(r, 5, f"=A{r}*Assumptions!$B$22*$B$4").number_format = cur0
        ws3.cell(r, 6, f"=Assumptions!$B$6+A{r}*Assumptions!$C$6").number_format = cur0
        ws3.cell(r, 7, f"=Assumptions!$B$7+A{r}*Assumptions!$C$7").number_format = cur0
        ws3.cell(r, 8, f"=Assumptions!$B$8+A{r}*Assumptions!$C$8").number_format = cur0

chart2 = LineChart()
chart2.title = "Cumulative 5-year cash TCO"
chart2.style = 10
chart2.y_axis.title = "Cumulative $"
chart2.x_axis.title = "Year"
chart2.height = 12
chart2.width = 18
data2 = Reference(ws3, min_col=2, min_row=6, max_col=8, max_row=12)
cats2 = Reference(ws3, min_col=1, min_row=7, max_row=12)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
ws3.add_chart(chart2, "A15")

ws3["A32"] = "Year-5 cash comparison by intensity"
ws3["A32"].font = section
for i, h in enumerate(
    ["Intensity", "5yr Rent", "5yr Lab_5k", "5yr Lab_15k", "Cash winner", "Rent - Lab5k"],
    1,
):
    ws3.cell(33, i, h)
style_header_row(ws3, 33, 6)

for i, brow in enumerate([19, 20, 21, 22]):
    r = 34 + i
    label = ["Sparse", "Moderate", "Heavy", "Industrial"][i]
    ws3.cell(r, 1, label)
    ws3.cell(r, 2, f"=5*Assumptions!B{brow}*$B$4").number_format = cur0
    ws3.cell(r, 3, "=Assumptions!B6+5*Assumptions!C6").number_format = cur0
    ws3.cell(r, 4, "=Assumptions!B7+5*Assumptions!C7").number_format = cur0
    ws3.cell(r, 5, f'=IF(B{r}<=C{r},"RENT","OWN Lab_5k")')
    ws3.cell(r, 6, f"=B{r}-C{r}").number_format = cur0

ws3["A39"] = (
    "Positive delta = rent costs MORE than Lab_5k over 5 years (ownership wins on cash)."
)
ws3["A39"].font = Font(name="Arial", italic=True, size=9, color="666666")

for col in range(1, 9):
    ws3.column_dimensions[get_column_letter(col)].width = 14
ws3.column_dimensions["A"].width = 16

# ========== ADJUSTED TCO ==========
ws4 = wb.create_sheet("Adjusted_TCO")
ws4["A1"] = "Adjusted TCO — cash + sovereignty − rental data risk"
ws4["A1"].font = title
ws4.merge_cells("A1:G1")
ws4["A2"] = (
    "Own gets annual sovereignty credit; rent gets annual data-risk charge. Both from Assumptions."
)
ws4["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

ws4["A4"] = "Sovereignty benefit $/yr (own)"
ws4["B4"] = "=Assumptions!B28"
ws4["B4"].number_format = cur0
ws4["A5"] = "Data-risk cost $/yr (rent)"
ws4["B5"] = "=Assumptions!B29"
ws4["B5"].number_format = cur0

for i, h in enumerate(
    [
        "Intensity",
        "5yr Rent cash",
        "5yr Rent + risk",
        "5yr Lab_5k cash",
        "5yr Lab − sov credit",
        "Adj winner",
        "Adj rent − adj own",
    ],
    1,
):
    ws4.cell(7, i, h)
style_header_row(ws4, 7, 7)

for i, brow in enumerate([19, 20, 21, 22]):
    r = 8 + i
    label = ["Sparse", "Moderate", "Heavy", "Industrial"][i]
    ws4.cell(r, 1, label)
    ws4.cell(
        r,
        2,
        f"=5*Assumptions!B{brow}*INDEX(Assumptions!D12:D15,Assumptions!B25-11)",
    ).number_format = cur0
    ws4.cell(r, 3, f"=B{r}+5*$B$5").number_format = cur0
    ws4.cell(r, 4, "=Assumptions!B6+5*Assumptions!C6").number_format = cur0
    ws4.cell(r, 5, f"=D{r}-5*$B$4").number_format = cur0
    ws4.cell(r, 6, f'=IF(C{r}<=E{r},"RENT","OWN")')
    ws4.cell(r, 7, f"=C{r}-E{r}").number_format = cur0

ws4["A13"] = (
    "Positive G = rent still more expensive after risk/sov → OWN preferred on adjusted basis."
)
ws4["A13"].font = Font(name="Arial", italic=True, size=9, color="666666")

chart3 = BarChart()
chart3.type = "col"
chart3.grouping = "clustered"
chart3.title = "5yr adjusted: Rent+risk vs Lab5k−sov"
chart3.y_axis.title = "$"
chart3.height = 10
chart3.width = 14
data3 = Reference(ws4, min_col=3, min_row=7, max_col=5, max_row=11)
cats3 = Reference(ws4, min_col=1, min_row=8, max_row=11)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws4.add_chart(chart3, "A15")

for col in range(1, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 16

# ========== SCENARIO GRID ==========
ws6 = wb.create_sheet("Scenario_Grid")
ws6["A1"] = "Full scenario grid — 5yr cash TCO ($)"
ws6["A1"].font = title
ws6["A3"] = "Runs/yr"
run_list = [4, 12, 24, 36, 52, 80, 120]
for i, rn in enumerate(run_list):
    c = ws6.cell(3, 2 + i, rn)
    c.font = blue
    c.fill = yellow

ws6["A4"] = "Option"
for i, rn in enumerate(run_list):
    cell = ws6.cell(4, 2 + i, f"{rn}/yr")
    cell.font = header
    cell.fill = navy
ws6["A4"].font = header
ws6["A4"].fill = navy

ws6["A5"] = "Rent 5yr"
ws6["A6"] = "Lab_5k 5yr"
ws6["A7"] = "Lab_15k 5yr"
ws6["A8"] = "Lab_25k 5yr"
ws6["A9"] = "Winner"

for i in range(len(run_list)):
    c = 2 + i
    col = get_column_letter(c)
    ws6.cell(
        5,
        c,
        f"=5*{col}3*INDEX(Assumptions!$D$12:$D$15,Assumptions!$B$25-11)",
    ).number_format = cur0
    ws6.cell(6, c, "=Assumptions!$B$6+5*Assumptions!$C$6").number_format = cur0
    ws6.cell(7, c, "=Assumptions!$B$7+5*Assumptions!$C$7").number_format = cur0
    ws6.cell(8, c, "=Assumptions!$B$8+5*Assumptions!$C$8").number_format = cur0
    ws6.cell(
        9,
        c,
        f'=IF({col}5=MIN({col}5:{col}8),"RENT",IF({col}6=MIN({col}5:{col}8),"L5k",'
        f'IF({col}7=MIN({col}5:{col}8),"L15k","L25k")))',
    )

ws6["A11"] = (
    "Winner = pure cash minimum. L15k/L25k rarely win cash-only unless runs or $/job are high."
)
ws6["A11"].font = Font(name="Arial", italic=True, size=9, color="666666")

for col in range(1, 10):
    ws6.column_dimensions[get_column_letter(col)].width = 12
ws6.column_dimensions["A"].width = 14

# ========== FINDINGS ==========
ws5 = wb.create_sheet("Executive_Findings", 1)
ws5["A1"] = "Executive findings — Compute cost-benefit (Sovereign Mirror)"
ws5["A1"].font = title
ws5.merge_cells("A1:C1")
ws5["A2"] = (
    "Decision support | Operator: build lab vs rent GSTD/cloud | Horizon: 5 years | "
    "Open model: recompute when yellow inputs change"
)
ws5["A2"].font = Font(name="Arial", size=10, color="666666")

ws5["A4"] = "ID"
ws5["B4"] = "Finding"
ws5["C4"] = "Detail"
style_header_row(ws5, 4, 3)

findings = [
    (
        "F1",
        "Cash break-even is intensity-driven",
        "At default ~$80/job (32B blended), Lab_5k needs ~62 runs to recover capex. "
        "That is ~15.6 yrs at Sparse (4/yr), ~2.6 yrs Moderate (24/yr), ~1.2 yrs Heavy (52/yr), "
        "~0.5 yrs Industrial (120/yr). Ownership is not always cheaper — it is cheaper when you iterate.",
    ),
    (
        "F2",
        "Sparse training favors rental / GSTD-class nets",
        "Few trains per year while learning scaffolds: pay-per-job dominates $5–25k capex. "
        "Capital stays liquid; silicon is not idle. GSTD competes with cloud on $/hr — not free.",
    ),
    (
        "F3",
        "Moderate-to-heavy iteration flips to OWN on cash alone",
        "Weekly / twice-monthly QLoRA (living sovereign mirror) tends to make Lab_5k cheaper "
        "over 5 years than renting equivalent jobs, before any sovereignty premium.",
    ),
    (
        "F4",
        "Lab_15k / Lab_25k need heavier use or non-cash value",
        "Larger labs clear pure cash break-even only at high run rates or expensive jobs (70B). "
        "Buy up for 32B+ headroom / multi-experiment — not status.",
    ),
    (
        "F5",
        "Adjusted TCO shifts frontier toward OWN",
        "Modest $/yr for (a) not uploading rope and (b) offline/no-token control moves several "
        "Moderate cases from RENT to OWN. Set shadow prices explicitly (Assumptions B28–B29).",
    ),
    (
        "F6",
        "GSTD is not free compute",
        "Decentralized nets still price GPU time. Prefer when cheaper AND capable of custom LoRA "
        "+ private datasets. Pure consumption = cloud with a token skin; running a node is different.",
    ),
    (
        "F7",
        "Optimal policy is hybrid",
        "(1) $0 craft; (2) $50–300 proof job; (3) Lab_5k when intensity approaches Moderate; "
        "(4) rent 70B bursts; (5) delay Lab_15/25 until 32B is routine. Minimizes regret under "
        "uncertain train frequency.",
    ),
    (
        "F8",
        "Dominant sensitivity is $/job",
        "If rental falls to $15/job, even Heavy use delays $25k justification. If $150+/job, "
        "Lab_5k wins quickly. Re-quote yellow cells; all curves recompute.",
    ),
]

for i, (fid, title_f, detail) in enumerate(findings):
    r = 5 + i
    ws5.cell(r, 1, fid).font = Font(name="Arial", bold=True)
    ws5.cell(r, 2, title_f).font = Font(name="Arial", bold=True, size=11)
    ws5.cell(r, 3, detail).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.row_dimensions[r].height = 58

ws5["A14"] = "RECOMMENDATION MATRIX"
ws5["A14"].font = section
for i, h in enumerate(["If expected runs/year…", "Primary", "Secondary"], 1):
    ws5.cell(15, i, h)
style_header_row(ws5, 15, 3)
recs = [
    ("≤ 6 (Sparse)", "RENT / GSTD for trains", "No lab until intensity proven"),
    ("12–30 (Moderate)", "Lab_5k within 12 months", "Rent for 70B bursts only"),
    ("40–60 (Heavy)", "Lab_5k now; plan 2nd GPU", "GSTD only if cheaper than power+time"),
    ("100+ (Industrial)", "Lab_15k–25k path", "Hybrid multi-node / supply-side GSTD"),
]
for i, row in enumerate(recs):
    for j, v in enumerate(row, 1):
        ws5.cell(16 + i, j, v)

ws5["A21"] = "DECISION RULE"
ws5["A21"].font = section
ws5["A22"] = (
    "OWN if: (Runs/yr × $/run × Years) + Risk$/yr×Years  >  Capex + Ops×Years − Sov$/yr×Years"
)
ws5["A22"].font = Font(name="Arial", bold=True, size=11)
ws5["A23"] = "Else RENT. Recompute when any yellow input changes. See BreakEven + TCO_Curves sheets."
ws5["A23"].font = Font(name="Arial", size=10, color="666666")

ws5["A25"] = "ILLUSTRATIVE POINT ESTIMATES (default assumptions, 32B job ~$80)"
ws5["A25"].font = section
ws5["A26"] = "Metric"
ws5["B26"] = "Value"
style_header_row(ws5, 26, 2)
points = [
    ("Lab_5k break-even runs @ $80/job", "62.5"),
    ("5yr Sparse rent (4×$80×5)", "$1,600"),
    ("5yr Moderate rent (24×$80×5)", "$9,600"),
    ("5yr Heavy rent (52×$80×5)", "$20,800"),
    ("5yr Lab_5k cash (5k+5×250)", "$6,250"),
    ("Cash winner Sparse", "RENT (save ~$4.6k)"),
    ("Cash winner Moderate", "OWN (save ~$3.4k)"),
    ("Cash winner Heavy", "OWN (save ~$14.6k)"),
]
for i, (m, v) in enumerate(points):
    ws5.cell(27 + i, 1, m)
    ws5.cell(27 + i, 2, v).font = Font(name="Arial", bold=True)

ws5.column_dimensions["A"].width = 48
ws5.column_dimensions["B"].width = 42
ws5.column_dimensions["C"].width = 78

out = "docs/compute_cost_curves.xlsx"
wb.save(out)
print("saved", out)
